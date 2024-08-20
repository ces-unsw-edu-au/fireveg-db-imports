import openpyxl
from datetime import datetime
import re
import copy

# Create field site records
def create_field_site_record(item,sw):
    site_label = item[sw['site_label']].value
    if site_label is not None and site_label != "Site":
        record={'site_label': site_label}
    
        for column in ('elevation','location_description', 'gps_uncertainty_m', 'gps_geom_description'):
            if column in sw.keys():
                val=item[sw[column]].value
                if val is not None and val not in ('na','NA'):
                    record[column] =  val
    
        if 'lons' in sw.keys():
            for xs in sw['lons']:
                xlon = item[xs].value
            for ys in sw['lats']:
                ylat = item[ys].value
            srid = 4326


        if 'xs' in sw.keys():
            for xs in sw['xs']:
                xlon = item[xs].value
            for ys in sw['ys']:
                ylat = item[ys].value

            if 'fixed_utm_zone' in sw.keys():
                utm_zone=sw['fixed_utm_zone']
            else:
                utm_zone=item[sw['utm_zone']].value
            if  utm_zone == 56:
                srid = 28356
            elif utm_zone == 55:
                srid = 28355
            elif utm_zone == 54:
                srid = 28354
            else:
                srid = False

   
        if srid and xlon and ylat:
            record['geom'] = "ST_GeomFromText('POINT({xlon} {ylat})', {srid})".format(xlon=xlon,ylat=ylat,srid=srid)

        return(record)

# Create field visit records
def create_field_visit_record(item,sw):
    site_label = item[sw['site_label']].value
    records = list()
    for k in sw['visit_date']:
        visit_date = item[k].value
        if site_label is not None and site_label != "Site":
            if isinstance(visit_date, datetime):
                record = {'visit_id': site_label, 'visit_date': visit_date}
                if 'survey' in sw.keys():
                    record['survey_name'] = sw['survey']
                for column in ('visit_description', 'mainobserver', 'observerlist','replicate_nr'):
                    if column in sw.keys():
                        val=item[sw[column]].value
                        if val is not None and val not in ('na','NA'):
                            if column=='observerlist':
                                val=val.split(',')
                            record[column] =  val
                records.append(record)
    return records

# Create fire history records
# This is a lower level function that will create a field sample record from an item (a row in the spreadsheet), using the dictionary or "switch" in col_dicts:

def create_fire_history_record(item,col_dicts):
    records=list()
    for sw in col_dicts:
        record=dict()
        comms=list()
        if item[sw['site_label']].value == 'Site':
            continue
        for k in sw.keys():
            vals=item[sw[k]].value
            if vals is not None:
                if k == 'fire_date':
                    if isinstance(vals,datetime):
                        record['fire_date']=str(vals.date())
                        record['earliest_date']=vals.date()
                        record['latest_date']=vals.date()
                    elif isinstance(vals,int):
                        record['fire_date']=str(vals)
                        if vals>0:
                            record['earliest_date']=datetime(vals,1,1).date()
                            record['latest_date']=datetime(vals,12,31).date()
                        else:
                            comms.append('Fire date is missing or empty')
                    elif vals.isnumeric():
                        record['fire_date']=str(vals)
                        record['earliest_date']=datetime(int(vals),1,1).date()
                        record['latest_date']=datetime(int(vals),12,31).date()
                    else:
                        record['fire_date']=vals
                        comms.append('Fire date given as: %s' % vals)
                        found=re.findall("[<>]",vals)
                        for i in found:
                            comms.append('max/min value given')
                            vals=vals.replace(i,"")
                        ws=vals.split("-")
                        if len(ws)==2:
                            if len(ws[0])==4 and ws[0].isnumeric():
                                record['earliest_date']=datetime(int(ws[0]),1,1).date()
                            if len(ws[1])==4 and ws[1].isnumeric():
                                record['latest_date']=datetime(int(ws[1]),12,31).date()
                            elif len(ws[1])==2 and ws[1].isnumeric():
                                fch2=ws[0][0:2]+ws[1]
                                record['latest_date']=datetime(int(fch2),12,31).date()
    
                            
                else:
                    record[k]=vals
        if len(comms)>0:
            record['notes'] = comms
        if len(record)>1:
            records.append(record)
    return records

## This is a lower level function that will create a field sample record from an `item` (a row in the spreadsheet), using the dictionary or "switch" in `sw`:

def create_field_sample_record(item,sw):
    visit_id=item[sw['visit_id']].value
    if visit_id is not None and visit_id not in ('Site Number'):
        if 'split_visit_id' in sw.keys():
            vid = visit_id.split("_")
            if not vid[2]:
                replicatenr = None
            else:
                replicatenr = int(vid[2])
            samplenr = vid[1]
            visit_id = vid[0]
        else:
            if 'replicate_nr' in sw.keys():
                replicatenr = item[sw['replicate_nr']].value
            elif 'fixed_replicate_nr' in sw.keys():
                replicatenr = sw['fixed_replicate_nr']
            else:
                replicatenr = None
            if 'sample_nr' in sw.keys():
                samplenr = item[sw['sample_nr']].value
            else:
                samplenr = None
        record={'visit_id': visit_id, 'replicate_nr': replicatenr, 'sample_nr': samplenr}
        if 'date' in sw.keys():
            visit_date = item[sw['date']].value 
            if isinstance(visit_date,datetime):
                record['visit_date'] = visit_date.date()
        return(record)


def create_quadrat_sample_record(item,sw,lookup,valid_seedbank,valid_organ):
    species = item[sw['species']].value
    visit_id =  item[sw['visit_id']].value
    comms=list()
    if 'sample_nr' in sw.keys():
        sample_nr=item[sw['sample_nr']].value
    else:
        sample_nr=1
    if 'replicate_nr' in sw.keys():
        replicate_nr = item[sw['replicate_nr']].value
    elif 'fixed_replicate_nr' in sw.keys():
        replicate_nr = sw['fixed_replicate_nr']
    else :
        replicate_nr = 1
            
    if 'split_visit_id' in sw.keys():
        if visit_id is not None:
            comms.append("visit_id originally recorded as %s" % visit_id)
            vid = visit_id.split("_")
            if not vid[2]:
                replicate_nr = None
            else:
                replicate_nr = int(vid[2])
            sample_nr = vid[1]
            visit_id = vid[0]
            
    if species is not None:
        record={'visit_id': visit_id, 'sample_nr': sample_nr,
                'species': species}
        if 'workbook' in sw.keys():
            comms.append("Imported from workbook %s using python script" % sw['workbook'])
        if 'worksheet' in sw.keys():
            comms.append("Imported from spreadsheet %s" % sw['worksheet'])
    
        if 'date' in sw.keys():
            visit_date = item[sw['date']].value
        else:
            visit_date = None
        
        if isinstance(visit_date,datetime):
            record['visit_date'] = visit_date.date()
        else:    
            p=filter(lambda n: n['visit_id'] == visit_id and  n['replicate_nr'] == replicate_nr, lookup)
            found=list(p)
            if len(found)==1 and 'visit_date' in found[0].keys():
                visit_date=found[0]['visit_date']
                if isinstance(visit_date,datetime):
                    record['visit_date'] = visit_date.date()
                    comms.append("visit date not provided, matched by replicate nr %s" % replicate_nr)
                else:
                    record['visit_date'] = visit_date
                    comms.append("matched by replicate nr %s, assuming date object" % replicate_nr)
            else:
                record['replicate_nr'] = replicate_nr
                comms.append("neither visit date nor replicate nr was matched ( replicate nr %s ), no date" % replicate_nr)

        if 'spcode' in sw.keys():
            spcode = item[sw['spcode']].value
            if (isinstance(spcode, str) and spcode.isnumeric()) or isinstance(spcode,int):
                record['species_code']=spcode
         
        for k in ('species_notes', 'resprout_organ', 'seedbank', 'adults_unburnt', 'resprouts_live', 'resprouts_died', 'resprouts_kill', 'resprouts_reproductive',
                  'recruits_live', 'recruits_reproductive', 'recruits_died','notes'):
            if k in sw.keys():
                vals=item[sw[k]].value
                if vals is not None and vals not in ('na','NA'):
                    if k == 'resprout_organ':
                        if vals in valid_organ:
                            record[k]=vals
                        elif vals.capitalize() in valid_organ:
                            record[k]=vals.capitalize()
                        else:
                            comms.append("resprout organ written as %s" % vals)
                    elif k == 'seedbank':
                        if vals in valid_seedbank:
                            record[k]=vals
                        elif vals.capitalize() in valid_seedbank:
                            record[k]=vals.capitalize()
                        else:
                            comms.append("seedbank written as %s" % vals)
                    elif k == 'notes':
                        if isinstance(vals,(int, float, complex)):
                            comms.append("Comment column included a numeric value of %s" % vals)
                        else:
                            comms.append(vals)
                    elif k in ('adults_unburnt', 'resprouts_live', 'resprouts_died', 'resprouts_kill', 'resprouts_reproductive',
                  'recruits_live', 'recruits_reproductive', 'recruits_died'):
                        if isinstance(vals,int):
                            record[k]=vals   
                        else:
                            comms.append("%s written as %s" % (k,vals))
                    else:
                        record[k]=vals        
        
        if len(comms)>0:
            record["comments"]=comms
        
        return(record)



### Functions to read records in a workbook
# We need a wrapping function to apply a lower level function (`create_record_function`) to all rows in a `worksheet` of the selected `workbook` using a dictionary `col_dictionary`, we add a `**kwargs` to pass additional arguments to the lower level function:

def import_records_from_workbook(filepath, workbook, worksheet, col_dictionary, create_record_function, **kwargs):
    wb = openpyxl.load_workbook(filepath / workbook, data_only=True)
    ws=wb[worksheet]
    row_count = ws.max_row+1
    records=list()
    for k in range(2,row_count):
        item=ws[k]
        record=create_record_function(item,col_dictionary,**kwargs)
        if record is not None:
            if type(record)==list:
                records.extend(record)
            elif type(record)==dict:
                records.append(record)
    return records

def read_fire_intensity(filepath,workbook,worksheet,col_definitions):
    wb = openpyxl.load_workbook(filepath / workbook,data_only=True)
    ws = wb[worksheet]
    triplet=('best','lower','upper')
    records=list()
    for row in range(2,ws.max_row+1):
        visitid=ws.cell(row,col_definitions['visit_id']).value
        if visitid is not None and visitid != 'Site':
            visitdate=ws.cell(row,col_definitions['visit_date']).value
            if isinstance(visitdate,datetime):
                visitdate=visitdate.date()
            elif visitdate is None:
                visitdate='NULL'
            else:
                visitdate=datetime.strptime(visitdate, '%d/%m/%Y').date()
            
            record={'visit_id': visitid,
                    'visit_date': visitdate,
                   'comment':list()}
            #print(record)
            for var in ('scorch height',
                        'tree foliage biomass consumed', 'shrub foliage biomass consumed', 'ground foliage biomass consumed',
                        'tree foliage scorch', 'shrub foliage scorch', 'herb foliage scorch',
                        'peat extent burnt','peat depth burnt'
                       ):
                if var in col_definitions.keys():
                    record1=copy.deepcopy(record)
                    record1['measured_var']=var
                    if var == 'scorch height':
                        record1['units']='m'
                    elif var == 'peat depth burnt':
                        record1['units']='cm'
                    else:
                        record1['units']='%'
                    for k in range(len(col_definitions[var])):
                        val=ws.cell(row,col_definitions[var][k]).value
                        if val is not None and val != 'NA':
                            if triplet[k]=='lower' and 'best' in record1.keys() and val > record1['best']:
                                record1['lower']=record1['best']
                                record1['comment'].append('lower bound given as %s but greater than best estimate' % val)
                            if triplet[k]=='upper' and 'best' in record1.keys() and val < record1['best']:
                                record1['comment'].append('upper bound given as %s but less than best estimate' % val)
                                record1['upper']=record1['best']
                            else:
                                record1[triplet[k]]=val
                    #print(record1)
                    if 'comment' in record1.keys() and len(record1['comment'])==0:
                        record1.pop('comment')
                    records.append(record1)
    return records

# Add raw measurements for a single variable
def read_twig_diameters(filepath,workbook,worksheet,col_definitions):
    wb = openpyxl.load_workbook(filepath / workbook,data_only=True)
    ws = wb[worksheet]
    records=list()
    for row in range(2,ws.max_row+1):
        visitid=ws.cell(row,col_definitions['visit_id']).value
        if visitid is not None and visitid != 'Site':
            visitdate=ws.cell(row,col_definitions['visit_date']).value
            if isinstance(visitdate,datetime):
                visitdate=visitdate.date()
            elif visitdate is None:
                visitdate='NULL'
            else:
                visitdate=datetime.strptime(visitdate, '%d/%m/%Y').date()
             
            for var in ('twig diameter',):
                if var in col_definitions.keys():
                    record={'visit_id': visitid,
                        'visit_date': visitdate,
                        'measured_variable': var,
                        'units':'mm'
                       }
                    for k in range(len(col_definitions[var])):
                        record1=copy.deepcopy(record)
                        val=ws.cell(row,col_definitions[var][k]).value
                        if val is not None and val != 'NA':
                            record1['single_value']=val
                            records.append(record1)
    return records

# I defined this function to read vegetation information from each worksheet.
def read_veg_classes(filepath,workbook,worksheet,col_definitions):
    wb = openpyxl.load_workbook(filepath / workbook,data_only=True)
    ws = wb[worksheet]
    records=list()
    for row in range(2,ws.max_row+1):
        visitid=ws.cell(row,col_definitions['visit_id']).value
        if visitid is not None and visitid != 'Site':
            visitdate=ws.cell(row,col_definitions['visit_date']).value
            if isinstance(visitdate,datetime):
                visitdate=visitdate.date()
            else:
                visitdate=datetime.strptime(visitdate, '%d/%m/%Y').date()
            vegclass=ws.cell(row,col_definitions['vegetation_class']).value
            vegformation=ws.cell(row,col_definitions['vegetation_formation']).value
            if vegclass=='Warm temperate rainforests':
                vegclass='Southern Warm Temperate Rainforests'
            if vegclass=='Littoral rainforest':
                vegclass='Littoral rainforests'
            if vegformation=='Rainforests':
                vegclass=vegclass.title()
            if vegformation in ('Blue Mountains Cool Wet Eucalypt Forest','Wet Sclerophyll Forests (Shrubby sub-formation)'):
                vegformation='Wet sclerophyll forests (Shrubby subformation)'
            if vegclass=='Southern Tableland Wet Sclerophyll Forests':
                vegformation='Wet sclerophyll forests (Grassy subformation)'
            if vegclass=='Montane wet sclerophyll forests':
                vegformation='Wet sclerophyll forests (Grassy subformation)'
                vegclass='Montane Wet Sclerophyll Forests'
            if vegclass=='Alpine bogs and fens':
                vegclass='Alpine Bogs and Fens'
            
            record={'visit_id': visitid,
            'visit_date': visitdate,
            'vegetation_formation':vegformation,
                'vegetation_class':vegclass}
            records.append(record)
    return records

def read_veg_structure(filepath,workbook,worksheet,col_definitions):
    wb = openpyxl.load_workbook(filepath / workbook,data_only=True)
    ws = wb[worksheet]
    triplet=('best','lower','upper')
    records=list()
    for row in range(2,ws.max_row+1):
        visitid=ws.cell(row,col_definitions['visit_id']).value
        if visitid is not None and visitid != 'Site':
            visitdate=ws.cell(row,col_definitions['visit_date']).value
            if isinstance(visitdate,datetime):
                visitdate=visitdate.date()
            else:
                visitdate=datetime.strptime(visitdate, '%d/%m/%Y').date()
            record={'visit_id': visitid,
            'visit_date': visitdate}
            
            stage=ws.cell(row,col_definitions['stage']).value
            if stage is not None:
                record['comment']=['Stage: %s' % stage,]
            stratum=ws.cell(row,col_definitions['stratum']).value
            for var in ('height','cover','scorch'):
                if var in col_definitions.keys():
                    record1=copy.deepcopy(record)
                    record1['measured_var']='stratum %s %s' % (stratum,var)
                    for k in range(len(col_definitions[var])):
                        val=ws.cell(row,col_definitions[var][k]).value
                        if val is not None and val != 'NA':
                            if triplet[k]=='lower' and 'best' in record1.keys() and val > record1['best']:
                                record1['lower']=record1['best']
                                record1['comment'].append('lower bound given as %s but greater than best estimate' % val)
                            if triplet[k]=='upper' and 'best' in record1.keys() and val < record1['best']:
                                record1['comment'].append('upper bound given as %s but less than best estimate' % val)
                                record1['upper']=record1['best']
                            else:
                                record1[triplet[k]]=val
                    records.append(record1)
    return records

